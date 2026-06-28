# app/services/budget/stages/ingest.py
#
# Stage 1 — Ingest
#
# Two sequential sub-steps:
#   A. Code reads the prior-year budget xlsx to extract line labels, sections,
#      and prior_year amounts from the rightmost all-caps BUDGET column.
#      Nothing about column selection is delegated to the AI.
#   B. AI reads only the PDF financial report to extract YTD actuals for each
#      of the labels provided from sub-step A.
#
# Results are merged: labels / sections / prior_year from A, ytd_actual from B.
#
# This is the only stage that calls an LLM; all subsequent stages are deterministic.

import base64
import io
from collections.abc import Callable
from pathlib import Path

from app.config.settings import settings
from app.services.budget.exceptions import IngestFailed
from app.services.budget.schema import (
    AIExtractionResult,
    BudgetSection,
    IngestedLine,
    IngestResult,
)

_PROMPT_PATH = Path(__file__).parent.parent / "prompts" / "ingest.txt"

# Ordered keyword pairs for section-header classification.
# Checked in order — first match wins.
_INCOME_SECTION_KEYWORDS: list[tuple[str, BudgetSection]] = [
    ("reserve", BudgetSection.REVENUE_RESERVES),
    ("income", BudgetSection.REVENUE_OPERATING),
    ("revenue", BudgetSection.REVENUE_OPERATING),
    ("operating", BudgetSection.REVENUE_OPERATING),
]
_EXPENSE_SECTION_KEYWORDS: list[tuple[str, BudgetSection]] = [
    ("admin", BudgetSection.ADMINISTRATION),
    ("maintenance", BudgetSection.MAINTENANCE),
    ("maint", BudgetSection.MAINTENANCE),
    ("util", BudgetSection.UTILITIES),
    ("other", BudgetSection.OTHER),
    ("general", BudgetSection.OTHER),
    ("reserve", BudgetSection.RESERVES),
    ("capital", BudgetSection.RESERVES),
]


def _find_budget_sheet(wb):
    for name in wb.sheetnames:
        if "budget" in name.lower():
            return wb[name]
    return wb.worksheets[0]


def _classify_section_header(text: str, past_income_total: bool) -> BudgetSection | None:
    t = text.strip().lower()
    keywords = _EXPENSE_SECTION_KEYWORDS if past_income_total else _INCOME_SECTION_KEYWORDS
    for kw, section in keywords:
        if kw in t:
            return section
    return None


def _detect_label_col(ws) -> int:
    """
    Return the 1-based column index that contains row labels (section headers,
    line-item names). Scans columns 1-4 and returns the leftmost one that has
    at least three cells containing alphabetic text in the first 40 data rows.
    This handles formats where column A is blank or contains numeric GL codes
    and the actual labels live in column B.
    """
    import re

    for col in range(1, min(ws.max_column + 1, 5)):
        letter_count = sum(
            1
            for row in range(2, min(ws.max_row + 1, 42))
            if (v := ws.cell(row=row, column=col).value) is not None
            and not isinstance(v, int | float)
            and re.search(r"[a-zA-Z]", str(v))
        )
        if letter_count >= 3:
            return col
    return 1


def _parse_excel_budget(xlsx_bytes: bytes, budget_year: int) -> list[dict]:
    """
    Read the prior-year budget xlsx and return one dict per line item:
        {label: str, section: BudgetSection, prior_year: float | None}

    Prior-year column detection (year-aware):
      1. Parse the 4-digit year from every header cell that contains "budget".
         The column whose year == budget_year - 1 is the prior-year source.
      2. If that column has fewer than 3 numeric values (blank), fall back to
         whichever budget column has the most data — including the proposed-year
         column — since the association may have entered values there by mistake.
         (render.py always writes to the structurally correct destination column,
         so reading from any column here causes no corruption.)
      3. No year match at all → use the richest non-rightmost column, else any.

    Section is inferred from section-header rows.
    """
    import re as _re

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = _find_budget_sheet(wb)

    # Auto-detect which column holds row labels. Some formats use column A;
    # others (GL-code column in A, label text in B) use column B.
    label_col = _detect_label_col(ws)

    # ── Prior-year column detection ────────────────────────────────────────────
    # Ingest is the DATA SOURCE: it reads from the column that currently holds
    # the adopted prior-year budget values.  That column was the PROPOSED slot
    # from the previous pipeline run and carries year = budget_year - 1.
    # (render.py writes to a DIFFERENT column — year = budget_year - 2 — so
    # the two stages are deliberately decoupled.)
    #
    # Title cells like "Morton Village ADOPTED Operating Budget" are excluded
    # by requiring either a 4-digit year OR a short phrase (<=3 words).
    prior_year_int = budget_year - 1
    budget_cols: list[tuple[int, int | None]] = []  # (col_idx, parsed_year | None)

    for check_row in range(1, min(ws.max_row + 1, 6)):
        found_any = False
        for col in range(1, ws.max_column + 1):
            raw = ws.cell(row=check_row, column=col).value
            if raw is None:
                continue
            s = str(raw).strip()
            if "budget" not in s.lower():
                continue
            m = _re.search(r"\b(20\d{2})\b", s)
            if m:  # primary: year-stamped columns only; excludes titles and section headers
                budget_cols.append((col, int(m.group(1))))
                found_any = True
        if found_any:
            break

    # Fallback: no year-stamped budget column found (rare legacy templates that
    # use bare "BUDGET" or two-word "XXXX BUDGET" headers with no 4-digit year).
    if not budget_cols:
        for check_row in range(1, min(ws.max_row + 1, 6)):
            for col in range(1, ws.max_column + 1):
                raw = ws.cell(row=check_row, column=col).value
                if raw is None:
                    continue
                parts = str(raw).strip().split()
                if parts and parts[-1].upper() == "BUDGET" and len(parts) <= 2:
                    budget_cols.append((col, None))
            if budget_cols:
                break

    def _val_count(col: int) -> int:
        return sum(
            1
            for r in range(2, min(ws.max_row + 1, 62))
            if isinstance(ws.cell(row=r, column=col).value, int | float)
        )

    prior_year_col: int | None = None

    if budget_cols:
        # 1. Year-exact: prefer the column labeled budget_year-1 (the recently-
        #    adopted proposed slot that now holds the prior-year values).
        prior_matches = [c for c, y in budget_cols if y == prior_year_int]
        if prior_matches:
            candidate = prior_matches[0]
            if _val_count(candidate) >= 3:
                prior_year_col = candidate  # populated → done
            else:
                # Year match exists but column is blank — the data must be in
                # another column (e.g. the association filled the wrong slot).
                best = max((c for c, _ in budget_cols), key=_val_count)
                prior_year_col = best if _val_count(best) > _val_count(candidate) else candidate

        # 2. No year-exact match — pick the richest column.  Prefer non-rightmost
        #    (the rightmost is the new proposed slot, likely blank or formula-only).
        if prior_year_col is None:
            rightmost = max(c for c, _ in budget_cols)
            non_proposed = [c for c, _ in budget_cols if c != rightmost]
            if non_proposed:
                best_np = max(non_proposed, key=_val_count)
                prior_year_col = (
                    best_np
                    if _val_count(best_np) > 0
                    else max((c for c, _ in budget_cols), key=_val_count)
                )
            else:
                prior_year_col = rightmost

    if prior_year_col is None:
        prior_year_col = label_col + 1  # last resort: first column after labels

    # Numeric data columns: every column except the label column.
    # Text values in the label column are filtered by isinstance anyway, but
    # excluding it keeps section-header detection clean.
    data_cols = [c for c in range(1, ws.max_column + 1) if c != label_col]

    lines: list[dict] = []
    current_section: BudgetSection = BudgetSection.REVENUE_OPERATING
    past_income_total = False
    seen_section_header = False  # skip preamble rows before the first real section

    for row in range(2, ws.max_row + 1):
        label_cell = ws.cell(row=row, column=label_col).value
        if label_cell is None:
            continue
        label = str(label_cell).strip()
        if not label:
            continue

        label_upper = label.upper()

        # Subtotal / total rows: skip, but use income-total to mark the transition.
        if label_upper.startswith("TOTAL") or label_upper.startswith("[SUBTOTAL]"):
            if not past_income_total and ("INCOME" in label_upper or "REVENUE" in label_upper):
                past_income_total = True
            continue

        # Collect numeric values from data columns.
        raw_data = [ws.cell(row=row, column=c).value for c in data_cols]
        numeric_vals = [v for v in raw_data if isinstance(v, int | float)]

        # A row is a section header when it has no numeric values, OR when all
        # numeric values are zero and the label matches a known section keyword.
        # The second case handles "RESERVES" rows that carry a stray 0.00 in a
        # previously rendered column, which would otherwise cause all reserve
        # line items to be misclassified under the preceding expense section.
        maybe_section = _classify_section_header(label, past_income_total)
        if not numeric_vals or (all(v == 0 for v in numeric_vals) and maybe_section is not None):
            if maybe_section is not None:
                current_section = maybe_section
                seen_section_header = True
            if not past_income_total and "expense" in label.lower():
                past_income_total = True
            continue

        # Skip rows that appear before the first real section header (preamble
        # info like per-unit assessment rates, date headers, etc.). These have
        # numeric values but are not budget line items; including them would send
        # misleading labels to the AI and inflate section subtotals.
        if not seen_section_header:
            continue

        # Line-item row: pull prior_year from the identified column.
        raw_val = ws.cell(row=row, column=prior_year_col).value
        prior_year = float(raw_val) if isinstance(raw_val, int | float) else None

        lines.append({"label": label, "section": current_section, "prior_year": prior_year})

    return lines


def _format_line_list(excel_lines: list[dict]) -> str:
    """Format the Excel-extracted line list as text for the AI prompt."""
    from collections import defaultdict

    by_section: dict[str, list[str]] = defaultdict(list)
    for line in excel_lines:
        by_section[line["section"].value].append(line["label"])

    section_order = [
        "REVENUE_OPERATING",
        "REVENUE_RESERVES",
        "ADMINISTRATION",
        "MAINTENANCE",
        "UTILITIES",
        "OTHER",
        "RESERVES",
    ]
    parts = ["BUDGET LINE ITEMS (match each to its YTD actual in the PDF):"]
    for sec in section_order:
        labels = by_section.get(sec, [])
        if labels:
            parts.append(f"\n{sec}:")
            for lbl in labels:
                parts.append(f"  - {lbl}")
    return "\n".join(parts)


def run(
    financial_report_bytes: bytes,
    financial_report_filename: str,
    prior_budget_bytes: bytes,
    prior_budget_filename: str,
    budget_year: int,
    on_step: Callable[[str], None] | None = None,
) -> IngestResult:
    """
    Stage 1: extract raw values from both input files.

    Sub-step A — code parses the xlsx to get labels, sections, and prior_year.
    Sub-step B — AI reads the PDF to get YTD actuals for each label.
    Results are merged into IngestResult.

    Raises IngestFailed if the xlsx cannot be parsed or if the LLM call fails.
    """
    # ── Sub-step A: deterministic Excel parsing ───────────────────────────────
    if on_step:
        on_step("reading_excel")
    try:
        excel_lines = _parse_excel_budget(prior_budget_bytes, budget_year)
    except Exception as exc:
        raise IngestFailed(f"Failed to parse prior-year budget xlsx: {exc}") from exc

    if not excel_lines:
        raise IngestFailed(
            "No line items found in the prior-year budget xlsx. "
            "Confirm the file has a recognizable budget column structure."
        )

    # ── DEBUG FLAG ────────────────────────────────────────────────────────────
    # Set to True to skip the AI call entirely. The pipeline will run with
    # ytd_actual=None for every line, so projected stays blank and proposed
    # falls back to prior_year. Use this to verify the Excel column-copy logic
    # is producing the right numbers before re-enabling the AI.
    _DEBUG_SKIP_AI = False
    if _DEBUG_SKIP_AI:
        merged = [
            IngestedLine(
                label=xl["label"],
                section=xl["section"],
                prior_year=xl["prior_year"],
                ytd_actual=None,
            )
            for xl in excel_lines
        ]
        return IngestResult(
            months_elapsed=6,
            lines=merged,
            missing_data=["[DEBUG] AI call skipped — verifying Excel column copy only"],
        )

    # ── Sub-step B: AI extracts YTD actuals from the PDF ─────────────────────
    if on_step:
        on_step("extracting")
    try:
        from anthropic import Anthropic
    except ImportError as e:
        raise IngestFailed("anthropic is not installed — run: pip install anthropic") from e

    client = Anthropic(api_key=settings.ANTHROPIC_API_KEY)
    system_prompt = _PROMPT_PATH.read_text()
    line_list_text = _format_line_list(excel_lines)

    pdf_b64 = base64.standard_b64encode(financial_report_bytes).decode()
    user_content: list[dict] = [
        {"type": "text", "text": f"INCOME AND EXPENSE REPORT ({financial_report_filename}):"},
        {
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": pdf_b64,
            },
        },
        {"type": "text", "text": line_list_text},
    ]

    try:
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=16384,
            system=system_prompt,
            tools=[
                {
                    "name": "ingest_result",
                    "description": "YTD actuals and supporting data extracted from the PDF.",
                    "input_schema": AIExtractionResult.model_json_schema(),
                }
            ],
            tool_choice={"type": "tool", "name": "ingest_result"},
            messages=[{"role": "user", "content": user_content}],
        )
    except Exception as exc:
        raise IngestFailed(f"LLM call failed: {exc}") from exc

    tool_block = next((b for b in response.content if b.type == "tool_use"), None)
    if tool_block is None:
        raise IngestFailed(
            "LLM did not return structured output — no tool_use block in response. "
            f"Stop reason: {response.stop_reason}"
        )

    try:
        ai_result = AIExtractionResult(**tool_block.input)
    except Exception as exc:
        raise IngestFailed(f"LLM returned invalid structured output: {exc}") from exc

    if not ai_result.months_elapsed or ai_result.months_elapsed <= 0:
        detail = (
            " The model reported: " + "; ".join(ai_result.missing_data)
            if ai_result.missing_data
            else ""
        )
        raise IngestFailed(
            "Could not determine months elapsed from the financial report. "
            "Confirm the report has a clear period header such as "
            "'For the Period Ending August 31, 2025'." + detail
        )

    # ── Merge: Excel (label/section/prior_year) + AI (ytd_actual/gl_account) ─
    # Build both an exact lookup and a case-insensitive fallback so that minor
    # AI casing differences ("Member Assessments" vs "Member ASSESSMENTS") still
    # match the Excel label rather than creating a phantom duplicate line.
    ai_by_label_exact: dict[str, object] = {line.label: line for line in ai_result.lines}
    ai_by_label_lower: dict[str, object] = {}
    for line in ai_result.lines:
        ai_by_label_lower.setdefault(line.label.strip().lower(), line)

    excel_label_set = {xl["label"] for xl in excel_lines}
    excel_label_set_lower = {xl["label"].strip().lower() for xl in excel_lines}

    merged: list[IngestedLine] = []

    for xl in excel_lines:
        ai_line = ai_by_label_exact.get(xl["label"]) or ai_by_label_lower.get(
            xl["label"].strip().lower()
        )
        merged.append(
            IngestedLine(
                label=xl["label"],
                section=xl["section"],
                prior_year=xl["prior_year"],
                gl_account=ai_line.gl_account if ai_line else None,
                ytd_actual=ai_line.ytd_actual if ai_line else None,
                annualization_review=ai_line.annualization_review if ai_line else False,
            )
        )

    # Lines the AI found in the PDF that aren't in the Excel (e.g., new revenue items).
    # Skip if the AI label is a case-variant of an existing Excel label — that means
    # the AI slightly mis-cased a provided label rather than returning a genuinely new line.
    for ai_line in ai_result.lines:
        if (
            ai_line.label not in excel_label_set
            and ai_line.label.strip().lower() not in excel_label_set_lower
        ):
            merged.append(
                IngestedLine(
                    label=ai_line.label,
                    section=ai_line.section,
                    prior_year=None,
                    gl_account=ai_line.gl_account,
                    ytd_actual=ai_line.ytd_actual,
                    annualization_review=ai_line.annualization_review,
                )
            )

    if not merged:
        raise IngestFailed(
            "No lines were produced. "
            "Confirm the prior-year budget xlsx and the financial report are valid."
        )

    return IngestResult(
        months_elapsed=ai_result.months_elapsed,
        lines=merged,
        reserve_balances=ai_result.reserve_balances,
        pdf_section_totals=ai_result.pdf_section_totals,
        missing_data=ai_result.missing_data,
    )
