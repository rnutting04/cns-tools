# app/services/budget/stages/render.py
#
# Stage 5 — Render
#
# Edits the previous year's xlsx workbook in place:
#   1. Detects which columns hold prior/projected/proposed/notes values
#   2. Updates column headers to the new budget year
#   3. Overwrites each data row's values from the validated BudgetOutput
#      - Data rows: raw numeric values
#      - Subtotal/total rows: live Excel formulas (=SUM(...) / difference)
#        referencing the rows just written, so the workbook re-totals itself
#        if a reviewer tweaks a number by hand.
#   4. Highlights recurring-but-variable line items (CPA/accounting fees,
#      annual corporate report, insurance, backflow inspection, tree
#      trimming, mulch) so a reviewer always double-checks them.
#
# Any sheets beyond the Budget tab are left untouched, so any charts or
# formatting the association has added are preserved automatically.

import io
import re
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from difflib import SequenceMatcher

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from app.services.budget import layout as layout_mod
from app.services.budget.exceptions import RenderFailed
from app.services.budget.schema import BudgetLine, BudgetOutput
from app.services.budget.workbook import normalize_to_xlsx

# Section groupings mirroring assemble.py's grand-total math, keyed by
# BudgetSection.value so they line up with subtotal_excel_row's keys.
_INCOME_SECTION_KEYS = {"REVENUE_OPERATING", "REVENUE_RESERVES"}
_OPERATING_EXPENSE_SECTION_KEYS = {"ADMINISTRATION", "MAINTENANCE", "UTILITIES", "OTHER"}
_EXPENSE_SECTION_KEYS = _OPERATING_EXPENSE_SECTION_KEYS | {"RESERVES"}

# Line items that are recurring but lumpy/variable (renewals, seasonal
# contracts) — always flagged for manual review regardless of AI-detected
# annualization concerns. Matched as case-insensitive substrings against the
# line label, since actual xlsx wording varies ("Accounting/CPA", "Insurance
# - Umbrella", etc.).
_ALWAYS_REVIEW_KEYWORDS = (
    "cpa",
    "annual corp",
    "corporate report",
    "insurance",
    "backflow",
    "tree trim",
    "mulch",
)


def _needs_review_flag(label: str) -> bool:
    key = label.lower()
    return any(kw in key for kw in _ALWAYS_REVIEW_KEYWORDS)


def _sum_formula(col: int, rows: list[int]) -> str | None:
    """
    Build a =SUM(...) formula over *rows*, collapsing consecutive runs into
    ranges — e.g. rows [38..47] become 'D38:D47' rather than a 10-cell list.
    Non-contiguous rows still fall back to comma-joined ranges/cells, e.g.
    'D5:D7,D9'.
    """
    if not rows:
        return None
    letter = get_column_letter(col)
    rows = sorted(rows)
    groups: list[str] = []
    start = prev = rows[0]
    for r in rows[1:]:
        if r == prev + 1:
            prev = r
            continue
        groups.append(f"{letter}{start}" if start == prev else f"{letter}{start}:{letter}{prev}")
        start = prev = r
    groups.append(f"{letter}{start}" if start == prev else f"{letter}{start}:{letter}{prev}")
    return f"=SUM({','.join(groups)})"


BLUE_FILL = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
# Amber: no YTD actual was found in the financial report for this line, so its
# projected cell is blank and its proposed value falls back to the prior year.
# Distinct from BLUE so a reviewer can tell "check this number" apart from
# "we had nothing to work from here".
AMBER_FILL = PatternFill(start_color="FFF0CC", end_color="FFF0CC", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)

# Words ignored when doing keyword-prefix matching on subtotal rows
_SUBTOTAL_STOPWORDS = {"total", "and", "the", "of", "expense", "expenses"}


@dataclass
class RenderResult:
    xlsx_bytes: bytes
    review_flags: list[str]


def _copy_row_style(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)


def _find_budget_sheet(wb):
    """Return the budget worksheet. Delegates to the shared layout module."""
    return layout_mod.pick_budget_sheet(wb)


def _detect_label_col(ws, header_row: int = 1) -> int:
    """Return the column holding row labels. Delegates to the shared layout module."""
    return layout_mod.detect_label_col(ws, header_row)


def _detect_column_positions(ws, budget_year: int, wb=None):
    """
    Locate the prior-year, projected, proposed and notes columns to WRITE into.

    Detection itself lives in app.services.budget.layout so that ingest and
    render can never disagree about a workbook's shape — they previously kept
    separate copies of this logic and shared the same blind spots (sheet found
    only by name, value columns found only via the literal word "budget").

    Returns (prior_col, projected_col, proposed_col, notes_col, header_row).
    """
    lay = layout_mod.build_layout(wb if wb is not None else ws.parent, budget_year)
    prior, projected, proposed, notes = layout_mod.render_columns(lay, budget_year)
    return prior, projected, proposed, notes, lay.header_row


def _row_shift(row: int, inserted_at: list[int]) -> int:
    """How far *row* moved down after rows were inserted above it."""
    return sum(1 for at in inserted_at if at <= row)


_HEADER_YEAR_RE = re.compile(r"\b20\d{2}\b")


def _set_header(ws, header_row: int, col: int, year: int, fallback: str) -> None:
    """
    Roll a column header forward to *year*, keeping the association's wording.

    Associations word these headers their own way — "2026 ADOPTED",
    "2026 PROPOSED BUDGET", "2026 Approved Budget", "Projected 2025" — so only
    the year is replaced. Rewriting the whole header to a house style would
    erase the sheet's own vocabulary for no benefit. The fallback is used only
    when the existing header carries no year to swap.
    """
    cell = ws.cell(row=header_row, column=col)
    current = cell.value
    if isinstance(current, str) and _HEADER_YEAR_RE.search(current):
        cell.value = _HEADER_YEAR_RE.sub(str(year), current, count=1)
    else:
        cell.value = fallback


_TITLE_RE = re.compile(r"operating\s+budget", re.I)
_MONTHS_RE = re.compile(r"january\s+1\s*,", re.I)


def _title_lines(association_name: str, budget_year: int) -> list[str]:
    """The three-line title block every association's budget should carry."""
    return [
        association_name.strip().upper(),
        f"{budget_year} APPROVED OPERATING BUDGET",
        f"JANUARY 1, {budget_year} - DECEMBER 31, {budget_year}",
    ]


def _ensure_title_block(ws, label_col: int, header_row: int, lines: list[str]) -> int:
    """
    Write the standard title block above the column headers.

    Returns the number of rows INSERTED, so the caller can shift every row index
    it already holds. Most workbooks already carry some form of title and are
    rewritten in place; those with none (headers on row 1) get rows inserted.

    Styling is copied from whatever title the sheet already had, so an
    association's own fonts and merges survive.
    """
    # Find an existing title: a row above the headers mentioning "operating
    # budget" (the middle line of the block).
    middle = None
    for row in range(1, header_row):
        for col in range(1, min(ws.max_column, 4) + 1):
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str) and _TITLE_RE.search(value):
                middle = row
                break
        if middle:
            break

    inserted = 0
    if middle is None:
        # No title at all. Make room: three title rows plus a blank separator.
        needed = 4 - (header_row - 1)
        if needed > 0:
            ws.insert_rows(1, needed)
            inserted = needed
            header_row += needed
        middle = max(2, header_row - 3)

    start = middle - 1
    if start < 1:
        ws.insert_rows(1, 1 - start + 1)
        inserted += 1 - start + 1
        start = 1
        middle += 1

    # Reuse the existing title's styling where there is one to copy.
    template = ws.cell(row=middle, column=label_col)
    for offset, text in enumerate(lines):
        cell = ws.cell(row=start + offset, column=label_col)
        cell.value = text
        if template.has_style:
            cell.font = copy(template.font)
            cell.alignment = copy(template.alignment)

    return inserted


def _rows_for_sections(
    subtotal_excel_row: dict[str, int],
    group_section: dict[str, str],
    section_keys: set[str],
) -> list[int]:
    """Excel rows of every subtotal that rolls up into *section_keys*.

    A section can own more than one subtotal row when the workbook splits it
    (MCP: BUILDING AND GROUNDS + MAINTENANCE), so grand totals sum all of them.
    """
    return sorted(
        row for group, row in subtotal_excel_row.items() if group_section.get(group) in section_keys
    )


def _find_row(
    label_to_rows: dict[str, list[int]],
    claimed: set[int],
    label: str,
    is_computed: bool,
) -> int | None:
    """
    Return the first unclaimed xlsx row matching *label*.

    Pass 1 — exact case-insensitive match.
    Pass 2 — keyword-prefix forward match: every significant word in the
              budget label prefix-matches some word in the xlsx row.
              "Total Administration" matches "TOTAL ADMIN."
    Pass 3 — keyword-prefix reverse match: every significant word in the
              xlsx row prefix-matches some word in the budget label.
              Handles xlsx rows that are shorter than the budget label, e.g.
              "TOTAL OPERATING AND" matching "Total Operating and Reserves".
    """
    key = label.strip().lower()

    # Pass 1: exact (case-insensitive)
    for row in label_to_rows.get(key, []):
        if row not in claimed:
            return row

    if not is_computed:
        return None

    sig_words = [w.strip(".,()[]/-+") for w in key.split() if w not in _SUBTOTAL_STOPWORDS]
    sig_words = [w for w in sig_words if len(w) >= 4]
    if not sig_words:
        return None

    # Pass 2: all budget sig-words must find a prefix match in the xlsx row.
    for xlsx_key, rows in label_to_rows.items():
        if "total" not in xlsx_key and "surplus" not in xlsx_key:
            continue
        xlsx_words = [w.strip(".,()[]/-+") for w in xlsx_key.split()]
        xlsx_words = [w for w in xlsx_words if len(w) >= 4]
        if not xlsx_words:
            continue
        if all(
            any(
                sw[: max(4, min(len(sw), len(xw)))] == xw[: max(4, min(len(sw), len(xw)))]
                for xw in xlsx_words
            )
            for sw in sig_words
        ):
            for row in rows:
                if row not in claimed:
                    return row

    # Pass 3: all xlsx sig-words must find a prefix match in the budget sig-words.
    # Catches abbreviated xlsx rows ("TOTAL OPERATING AND") that are a strict
    # subset of the full budget label ("Total Operating and Reserves").
    for xlsx_key, rows in label_to_rows.items():
        if "total" not in xlsx_key and "surplus" not in xlsx_key:
            continue
        xlsx_sig = [w.strip(".,()[]/-+") for w in xlsx_key.split() if w not in _SUBTOTAL_STOPWORDS]
        xlsx_sig = [w for w in xlsx_sig if len(w) >= 4]
        if not xlsx_sig:
            continue
        if all(
            any(
                xs[: max(4, min(len(xs), len(sw)))] == sw[: max(4, min(len(xs), len(sw)))]
                for sw in sig_words
            )
            for xs in xlsx_sig
        ):
            for row in rows:
                if row not in claimed:
                    return row

    # Pass 4: closest overall match among unclaimed TOTAL rows.
    #
    # Prefix matching cannot bridge an abbreviation or a typo, and real
    # workbooks have both: "Total Building and Grounds" must find the sheet's
    # "TOTAL BLDG & GROUNDS", and "Total Utilities" must find "TOTAL UTILITES".
    # Left unmatched, those subtotal rows keep no formula and the grand total
    # silently omits them. The threshold is deliberately high so unrelated
    # sections never collide.
    best_row, best_score = None, 0.0
    for xlsx_key, rows in label_to_rows.items():
        if "total" not in xlsx_key and "surplus" not in xlsx_key:
            continue
        unclaimed = [r for r in rows if r not in claimed]
        if not unclaimed:
            continue
        score = SequenceMatcher(None, key, xlsx_key).ratio()
        if score > best_score:
            best_row, best_score = unclaimed[0], score

    return best_row if best_score >= 0.72 else None


def run(budget: BudgetOutput, review_flags: list[str], prev_year_xlsx_bytes: bytes) -> RenderResult:
    """
    Stage 5: edit the previous year's xlsx with new budget values.

    Loads the workbook, locates the Budget sheet, detects column positions
    from the header row, shifts all headers to the new year, and writes
    prior/projected/proposed values for every line in BudgetOutput.

    Data rows receive raw numeric values. Subtotal and grand-total rows
    receive live =SUM(...)/difference formulas built from the exact rows
    just written this run — never formulas carried over from the prior
    workbook — so there's no risk of the #REF! errors that stale
    prior-year references used to cause after columns shift each cycle.

    Raises RenderFailed on any openpyxl error.
    """
    try:
        # Legacy .xls inputs are converted up front so every stage downstream
        # sees the same openpyxl workbook.
        prev_year_xlsx_bytes = normalize_to_xlsx(prev_year_xlsx_bytes)
        wb = openpyxl.load_workbook(io.BytesIO(prev_year_xlsx_bytes))

        # One layout pass drives every structural decision below — the same pass
        # ingest made, so both stages target the same sheet and columns. Detecting
        # the header row first also means the label-column scan starts below the
        # headers instead of guessing row 1.
        lay = layout_mod.build_layout(wb, budget.budget_year)
        ws = wb[lay.sheet_title]

        # Standard title block, written before anything else so that any rows it
        # inserts are absorbed once, here, rather than invalidating row indexes
        # computed later. openpyxl does not rewrite formula references when rows
        # shift, so this cannot be done after the =SUM formulas are written.
        title_offset = _ensure_title_block(
            ws,
            lay.label_col,
            lay.header_row,
            _title_lines(budget.association_name, budget.budget_year),
        )
        if title_offset:
            lay.header_row += title_offset
            lay.data_start_row += title_offset
            lay.section_rows = {r + title_offset: s for r, s in lay.section_rows.items()}
            lay.section_source_labels = {
                r + title_offset: v for r, v in lay.section_source_labels.items()
            }

        header_row = lay.header_row
        # Column A may hold GL codes with the real labels in column B; using the
        # detected column keeps label_to_rows keyed by text, not GL numbers.
        label_col = lay.label_col
        prior_col, projected_col, proposed_col, notes_col = layout_mod.render_columns(
            lay, budget.budget_year
        )

        # Load a second copy with data_only=True so we can read cached formula
        # results from the assessment preamble without destroying the formulas.
        wb_data = openpyxl.load_workbook(io.BytesIO(prev_year_xlsx_bytes), data_only=True)
        ws_data = wb_data[lay.sheet_title]
        if title_offset:
            # Keep the data-only copy row-aligned with the edited sheet.
            ws_data.insert_rows(1, title_offset)

        # Roll all three column headers forward a year, INCLUDING the proposed
        # column. Its values and formulas are still left alone — that column
        # belongs to the manager, who types the new year's numbers into it — but
        # its heading has to say the year those numbers are for, otherwise the
        # sheet ends up with two columns both labelled last year.
        prior_year_int = budget.budget_year - 1
        _set_header(ws, header_row, prior_col, prior_year_int, f"{prior_year_int} Budget")
        if projected_col is not None:
            _set_header(
                ws, header_row, projected_col, prior_year_int, f"Projected {prior_year_int}"
            )
        if proposed_col is not None:
            _set_header(
                ws, header_row, proposed_col, budget.budget_year, f"{budget.budget_year} BUDGET"
            )

        data_start = header_row + 1

        # Detect the preamble boundary: rows before the first recognized budget
        # section header (INCOME, OPERATING, etc.) are assessment tables that
        # must be handled separately from the main budget line-item loop.
        # Rows containing "assessment" are explicitly excluded so sub-headers
        # like "Quarterly Assessment" don't terminate preamble detection early.
        _SECTION_KW = frozenset(
            {
                "income",
                "revenue",
                "operating",
                "administration",
                "maintenance",
                "expense",
                "expenses",
                "utility",
                "utilities",
            }
        )
        preamble_end = data_start  # default: no preamble
        for _pr in range(data_start, ws.max_row + 1):
            _pv = ws.cell(row=_pr, column=label_col).value
            if _pv is None:
                continue
            _pt = str(_pv).strip().lower()
            if "assessment" in _pt:
                continue
            if any(kw in _pt for kw in _SECTION_KW):
                # Check only budget data columns for numeric values — GL code
                # columns (which contain integers like 6010) would otherwise
                # cause every row to be classified as "has numeric values".
                _has_nums = any(
                    isinstance(ws_data.cell(row=_pr, column=c).value, int | float)
                    for c in [prior_col, projected_col, proposed_col]
                    if c is not None
                )
                if not _has_nums:
                    preamble_end = _pr
                    break

        # Build label → row index from column A (lowercase, stripped).
        # Start from the row after the detected header so header text is never
        # treated as a line-item label.
        # Normalize in-cell newlines so "TOTAL OPERATING AND\nRESERVES" matches
        # the budget line label "Total Operating and Reserves".
        label_to_rows: dict[str, list[int]] = {}
        for row in range(data_start, ws.max_row + 1):
            val = ws.cell(row=row, column=label_col).value
            if val is not None:
                key = str(val).strip().lower().replace("\n", " ")
                label_to_rows.setdefault(key, []).append(row)

        # ── Pre-phase: insert rows for lines not found in the Excel ───────────
        # Lines the AI found in the PDF that weren't in the original workbook
        # (prior_year=None, so the left column will be blank/zero). We insert
        # them before their section's subtotal so the main loop writes them
        # normally. Processing top-to-bottom with a cumulative shift counter
        # avoids stale row references after each insertion.

        _subtotal_label_for_section = {
            (line.subtotal_group or line.section.value): line.label
            for line in budget.lines
            if line.is_computed and line.code.startswith("subtotal_")
        }

        # Determine which data lines are missing from the workbook.
        _inserted_at: list[int] = []
        _pre_claimed: set[int] = set()
        _unfound_by_section: dict[str, list[BudgetLine]] = defaultdict(list)
        for _line in budget.lines:
            if _line.is_computed:
                continue
            _r = _find_row(label_to_rows, _pre_claimed, _line.label, False)
            if _r is not None:
                _pre_claimed.add(_r)
            else:
                _unfound_by_section[_line.subtotal_group or _line.section.value].append(_line)

        if _unfound_by_section:
            # Locate each section's subtotal row.
            _st_claimed: set[int] = set()
            _subtotal_pos: dict[str, int] = {}
            for _sec, _lbl in _subtotal_label_for_section.items():
                _r = _find_row(label_to_rows, _st_claimed, _lbl, True)
                if _r is not None:
                    _subtotal_pos[_sec] = _r
                    _st_claimed.add(_r)

            # Insert rows top-to-bottom so earlier insertions don't shift later positions.
            _ordered_secs = sorted(
                _unfound_by_section.keys(),
                key=lambda s: _subtotal_pos.get(s, 9999),
            )
            _shift = 0
            for _sec in _ordered_secs:
                _new_lines = _unfound_by_section[_sec]
                _base = _subtotal_pos.get(_sec)
                if _base is None:
                    continue
                _insert_at = _base + _shift

                # Skip blank spacing rows above the subtotal so the new item
                # lands adjacent to the last real data row, not in the gap.
                while _insert_at > data_start + 1:
                    _prev_val = ws.cell(row=_insert_at - 1, column=label_col).value
                    if _prev_val is not None and str(_prev_val).strip():
                        break
                    _insert_at -= 1

                # Find a real line-item row for style: scan backwards past blanks
                # and subtotal rows to get a normal data-row style.
                _style_template = _insert_at - 1
                while _style_template > data_start:
                    _tmpl_val = ws.cell(row=_style_template, column=label_col).value
                    if _tmpl_val is not None and str(_tmpl_val).strip():
                        _tu = str(_tmpl_val).strip().upper()
                        if not _tu.startswith("TOTAL") and not _tu.startswith("[SUBTOTAL]"):
                            break
                    _style_template -= 1

                for _i, _nl in enumerate(_new_lines):
                    ws.insert_rows(_insert_at + _i)
                    _copy_row_style(ws, _style_template, _insert_at + _i)
                    ws.cell(row=_insert_at + _i, column=label_col).value = _nl.label
                    _inserted_at.append(_insert_at + _i)
                _shift += len(_new_lines)

            # Rebuild the label index so the main loop finds inserted rows.
            label_to_rows = {}
            for _r in range(data_start, ws.max_row + 1):
                _v = ws.cell(row=_r, column=label_col).value
                if _v is not None:
                    _k = str(_v).strip().lower().replace("\n", " ")
                    label_to_rows.setdefault(_k, []).append(_r)

        # The pipeline writes to exactly two detected columns:
        #   prior_col     ← prior_year (the adopted budget from last cycle)
        #   projected_col ← projected  (annualised YTD actuals)
        # proposed_col is left ENTIRELY alone — values, formulas and header. It
        # belongs to the manager, who fills in the new year's numbers by hand
        # using the two columns this stage writes. Blanking it (as this used to)
        # threw away the figures they start from.
        #
        # Preamble rows (assessment tables before the first section header) are
        # excluded from the wipe so their formulas survive.
        active_cols = [c for c in [prior_col, projected_col] if c is not None]

        for _wipe_col in active_cols:
            for _r in range(preamble_end, ws.max_row + 1):
                ws.cell(row=_r, column=_wipe_col).value = None

        # Assessment preamble: roll the old proposed_col value (computed from
        # the formula via the data-only workbook) into prior_col and projected_col
        # so the two historical columns carry last cycle's assessment amounts.
        # proposed_col is left untouched — its formulas auto-recalculate for the
        # new budget year when Excel opens the file.
        if preamble_end > data_start:
            for _pr in range(data_start, preamble_end):
                _pval = ws_data.cell(row=_pr, column=proposed_col).value
                if not isinstance(_pval, int | float):
                    continue
                ws.cell(row=_pr, column=prior_col).value = _pval
                if projected_col is not None:
                    ws.cell(row=_pr, column=projected_col).value = _pval

        def _raw_vals(line):
            """Map col → value for a line (data rows and computed rows alike)."""
            m: dict[int, object] = {prior_col: line.prior_year}
            if projected_col is not None:
                m[projected_col] = line.projected
            return m

        def _write_cols(row: int, col_formula_or_val: dict[int, object], fill=NO_FILL):
            for col, val in col_formula_or_val.items():
                cell = ws.cell(row=row, column=col)
                cell.value = val
                cell.fill = fill

        # --- tracking state ---
        claimed: set[int] = set()
        # Keyed by SUBTOTAL GROUP, not by section: a workbook that keeps
        # "BUILDING AND GROUNDS" separate from "MAINTENANCE" has a printed
        # subtotal for each, and each must receive its own =SUM over its own
        # rows rather than one merged figure.
        section_item_rows: dict[str, list[int]] = defaultdict(list)
        subtotal_excel_row: dict[str, int] = {}  # group → excel row of its subtotal
        group_section: dict[str, str] = {}  # group → BudgetSection value it rolls up to
        special_row: dict[
            str, int
        ] = {}  # code → excel row, for grand totals that reference each other

        # Rows shift when new line items are inserted above them, so the
        # layout's recorded subtotal rows are re-based by that shift.
        _sheet_subtotal_rows = {
            group: (r + title_offset + _row_shift(r + title_offset, _inserted_at))
            for group, r in budget.subtotal_rows.items()
        }

        for line in budget.lines:
            # A subtotal writes into the row the layout parser identified for its
            # group. Matching by label text is only a fallback: association
            # workbooks spell these rows every possible way ("TOTAL BLDG &
            # GROUNDS", "TOTAL UTILITES", and a bare "TOTAL OPERATING" used for
            # an expense block), and guessing from text does not scale.
            row = None
            if line.is_computed and line.subtotal_group:
                candidate = _sheet_subtotal_rows.get(line.subtotal_group)
                if candidate is not None and candidate not in claimed:
                    row = candidate
            if row is None:
                row = _find_row(label_to_rows, claimed, line.label, line.is_computed)
            if row is None:
                continue
            claimed.add(row)

            section_key = line.section.value

            if not line.is_computed:
                # ── Data row: write prior_year and projected ──────────────────
                raw = _raw_vals(line)
                flagged = _needs_review_flag(line.label)
                # A line with no projected value means the report had no YTD
                # actual to annualize — most often the label in the workbook does
                # not appear in the report under that name. The cell would simply
                # render blank, so make it visible instead.
                unmatched = line.projected is None
                fill = AMBER_FILL if unmatched else (BLUE_FILL if flagged else NO_FILL)
                for col in active_cols:
                    cell = ws.cell(row=row, column=col)
                    cell.value = raw.get(col)
                    cell.fill = fill
                ws.cell(row=row, column=label_col).fill = fill
                if proposed_col is not None:
                    ws.cell(row=row, column=proposed_col).fill = fill
                if notes_col is not None:
                    note = (line.note or "").strip()
                    if flagged:
                        tag = "Recurring but variable cost — verify against current contract/policy"
                        note = f"{note} — {tag}" if note else tag
                    ws.cell(row=row, column=notes_col).value = note
                    ws.cell(row=row, column=notes_col).fill = fill
                if unmatched:
                    review_flags.append(
                        f'"{line.label}" (row {row}) — no matching line found in the financial '
                        f"report, so there is no projected figure. The proposed value falls back "
                        f"to last year's budget; enter a figure manually if that is not right."
                    )
                elif flagged:
                    review_flags.append(
                        f'"{line.label}" (row {row}) flagged for manual review — recurring lumpy/variable cost'
                    )
                group_key = line.subtotal_group or section_key
                group_section[group_key] = section_key
                section_item_rows[group_key].append(row)

            elif line.code.startswith("subtotal_"):
                # ── Subtotal row: =SUM(...) over this GROUP's item rows ────────
                group_key = line.subtotal_group or section_key
                group_section[group_key] = section_key
                subtotal_excel_row[group_key] = row
                raw = _raw_vals(line)
                formulas = {
                    col: _sum_formula(col, section_item_rows.get(group_key, [])) or raw.get(col)
                    for col in active_cols
                }
                _write_cols(row, formulas)

            elif line.code == "total_income":
                sec_rows = _rows_for_sections(
                    subtotal_excel_row, group_section, _INCOME_SECTION_KEYS
                )
                raw = _raw_vals(line)
                formulas = {col: _sum_formula(col, sec_rows) or raw.get(col) for col in active_cols}
                _write_cols(row, formulas)
                special_row["total_income"] = row

            elif line.code == "total_operating":
                sec_rows = _rows_for_sections(
                    subtotal_excel_row, group_section, _OPERATING_EXPENSE_SECTION_KEYS
                )
                raw = _raw_vals(line)
                formulas = {col: _sum_formula(col, sec_rows) or raw.get(col) for col in active_cols}
                _write_cols(row, formulas)

            elif line.code == "total_operating_and_reserves":
                sec_rows = _rows_for_sections(
                    subtotal_excel_row, group_section, _EXPENSE_SECTION_KEYS
                )
                raw = _raw_vals(line)
                formulas = {col: _sum_formula(col, sec_rows) or raw.get(col) for col in active_cols}
                _write_cols(row, formulas)
                special_row["total_operating_and_reserves"] = row

            elif line.code == "surplus":
                ti_row = special_row.get("total_income")
                te_row = special_row.get("total_operating_and_reserves")
                raw = _raw_vals(line)
                formulas = {}
                for col in active_cols:
                    if ti_row is not None and te_row is not None:
                        letter = get_column_letter(col)
                        formulas[col] = f"={letter}{ti_row}-{letter}{te_row}"
                    else:
                        formulas[col] = raw.get(col)
                _write_cols(row, formulas)

            else:
                # Any other computed row type: fall back to a plain number.
                _write_cols(row, _raw_vals(line))

        # ── Post-pass: flag rows within section ranges that weren't matched ──────
        # These rows retain their existing values (intentional — they may be
        # valid budget items with a slightly different label). Just note them
        # so the reviewer can verify nothing was missed.
        for section_key, item_rows in section_item_rows.items():
            if not item_rows:
                continue
            first_row = min(item_rows)
            last_row = subtotal_excel_row.get(section_key, max(item_rows) + 1)
            for r in range(first_row, last_row):
                if r in claimed:
                    continue
                label_val = ws.cell(row=r, column=label_col).value
                if label_val is None:
                    continue
                label_str = str(label_val).strip()
                if not label_str:
                    continue
                review_flags.append(
                    f'Unmatched row in {section_key}: "{label_str}"'
                    f" (row {r}) — verify this line is accounted for"
                )

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

    except RenderFailed:
        raise
    except Exception as exc:
        raise RenderFailed(f"XLSX render failed: {exc}") from exc

    return RenderResult(xlsx_bytes=xlsx_bytes, review_flags=review_flags)
