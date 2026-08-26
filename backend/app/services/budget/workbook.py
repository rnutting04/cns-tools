# app/services/budget/workbook.py
#
# One loader for every prior-budget workbook the pipeline accepts.
#
# New association files are .xlsx. Older archives are still legacy binary .xls,
# which openpyxl cannot open at all — it raises BadZipFile or
# "File contains no valid workbook part", both of which surfaced as an opaque
# `ingest_failed`. This module converts .xls to .xlsx up front so every stage
# downstream sees an openpyxl workbook and nothing else has to care.
#
# The conversion carries values, number formats, bold/size, column widths and
# merges — enough for the render stage to write a presentable workbook back.
# It is deliberately not a full-fidelity converter; .xls is a legacy path.

import io

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# xlrd border line-style index → openpyxl side style name.
_BORDER_STYLES = {
    1: "thin",
    2: "medium",
    3: "dashed",
    4: "dotted",
    5: "thick",
    6: "double",
    7: "hair",
    8: "mediumDashed",
    9: "dashDot",
    10: "mediumDashDot",
    11: "dashDotDot",
    12: "mediumDashDotDot",
    13: "slantDashDot",
}

_H_ALIGN = {1: "left", 2: "center", 3: "right", 4: "fill", 5: "justify", 6: "centerContinuous"}
_V_ALIGN = {0: "top", 1: "center", 2: "bottom", 3: "justify"}


def _rgb(book, colour_index):
    """Resolve an xlrd colour index to an 'RRGGBB' string, or None."""
    if colour_index is None:
        return None
    try:
        rgb = book.colour_map.get(colour_index)
    except AttributeError:
        return None
    if not rgb:
        return None
    r, g, b = rgb[:3]
    return f"{r:02X}{g:02X}{b:02X}"


def _side(book, line_style, colour_index):
    style = _BORDER_STYLES.get(line_style)
    if not style:
        return Side()
    colour = _rgb(book, colour_index)
    return Side(style=style, color=colour) if colour else Side(style=style)


def _is_xls(data: bytes, filename: str) -> bool:
    """Legacy .xls files start with the OLE2 compound-document magic number."""
    return data[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" or filename.lower().endswith(".xls")


def xls_to_xlsx(data: bytes) -> bytes:
    """Convert a legacy binary .xls into .xlsx bytes."""
    try:
        import xlrd
    except ImportError as e:  # pragma: no cover - dependency is pinned
        raise RuntimeError("Reading legacy .xls files requires xlrd — run: pip install xlrd") from e

    try:
        src = xlrd.open_workbook(file_contents=data, formatting_info=True)
        has_formatting = True
    except NotImplementedError:
        # Some .xls variants (notably files written by non-Excel tools) do not
        # carry the formatting records xlrd needs. Values still convert fine.
        src = xlrd.open_workbook(file_contents=data)
        has_formatting = False

    out = openpyxl.Workbook()
    out.remove(out.active)

    for sheet in src.sheets():
        ws = out.create_sheet(title=(sheet.name or "Sheet")[:31])

        # Whether the sheet shows its cell grid is a per-sheet setting, and these
        # budget templates routinely turn it OFF and rule the table with explicit
        # borders instead. openpyxl defaults new sheets to gridlines ON, so
        # without carrying this across the output looked nothing like the source.
        show_grid = getattr(sheet, "show_grid_lines", None)
        if show_grid is not None:
            ws.sheet_view.showGridLines = bool(show_grid)

        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    continue
                dst = ws.cell(row=r + 1, column=c + 1)
                dst.value = cell.value

                if not has_formatting:
                    continue
                try:
                    xf = src.xf_list[sheet.cell_xf_index(r, c)]
                except (IndexError, AttributeError):
                    continue

                fmt = src.format_map.get(xf.format_key)
                if fmt is not None and fmt.format_str and fmt.format_str != "General":
                    dst.number_format = fmt.format_str

                # Font: name, size, weight, style, underline and colour.
                try:
                    f = src.font_list[xf.font_index]
                    dst.font = Font(
                        name=getattr(f, "name", None) or None,
                        bold=bool(f.bold),
                        italic=bool(f.italic),
                        underline="single" if getattr(f, "underline_type", 0) else None,
                        size=(f.height / 20) if f.height else None,
                        color=_rgb(src, getattr(f, "colour_index", None)),
                    )
                except (IndexError, AttributeError):
                    pass

                # Borders — the most visible loss when these are dropped, since
                # budget sheets rely on ruled lines to separate sections.
                try:
                    b = xf.border
                    dst.border = Border(
                        top=_side(src, b.top_line_style, b.top_colour_index),
                        bottom=_side(src, b.bottom_line_style, b.bottom_colour_index),
                        left=_side(src, b.left_line_style, b.left_colour_index),
                        right=_side(src, b.right_line_style, b.right_colour_index),
                    )
                except AttributeError:
                    pass

                # Cell shading, including white. A white fill is not a no-op: it
                # paints over the gridline beneath it. Copying it verbatim is
                # correct precisely because the sheet's own gridline setting is
                # carried across too, so the cell renders the same either way.
                try:
                    bg = xf.background
                    if bg.fill_pattern:
                        fg_rgb = _rgb(src, bg.pattern_colour_index)
                        bg_rgb = _rgb(src, bg.background_colour_index)
                        if fg_rgb:
                            dst.fill = PatternFill(
                                fill_type="solid",
                                start_color=fg_rgb,
                                end_color=bg_rgb or fg_rgb,
                            )
                except AttributeError:
                    pass

                try:
                    al = xf.alignment
                    dst.alignment = Alignment(
                        horizontal=_H_ALIGN.get(al.hor_align),
                        vertical=_V_ALIGN.get(al.vert_align),
                        wrap_text=bool(al.text_wrapped),
                        indent=getattr(al, "indent_level", 0) or 0,
                    )
                except AttributeError:
                    pass

        # Column widths.
        if has_formatting and getattr(sheet, "colinfo_map", None):
            for idx, info in sheet.colinfo_map.items():
                if idx < 16384:
                    letter = get_column_letter(idx + 1)
                    ws.column_dimensions[letter].width = info.width / 256
                    if info.hidden:
                        ws.column_dimensions[letter].hidden = True

        # Row heights — without these every row snaps to default and the sheet
        # reflows, which reads as "all the formatting is gone".
        if has_formatting and getattr(sheet, "rowinfo_map", None):
            for idx, info in sheet.rowinfo_map.items():
                if info.height is not None and not info.height_mismatch:
                    continue
                if info.height:
                    ws.row_dimensions[idx + 1].height = info.height / 20

        # Merged ranges.
        for rlo, rhi, clo, chi in getattr(sheet, "merged_cells", []) or []:
            ws.merge_cells(start_row=rlo + 1, start_column=clo + 1, end_row=rhi, end_column=chi)

    buf = io.BytesIO()
    out.save(buf)
    return buf.getvalue()


def normalize_to_xlsx(data: bytes, filename: str = "") -> bytes:
    """Return xlsx bytes for *data*, converting from legacy .xls when needed."""
    return xls_to_xlsx(data) if _is_xls(data, filename) else data


def load_workbook_any(data: bytes, filename: str = "", *, data_only: bool = True):
    """Load any supported workbook format as an openpyxl Workbook."""
    return openpyxl.load_workbook(
        io.BytesIO(normalize_to_xlsx(data, filename)), data_only=data_only
    )
