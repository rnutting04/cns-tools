"""Unit tests for workbook layout detection, the confidence gate, and reserve parsing."""

import io

import openpyxl
import pytest

from app.services.budget.layout import (
    build_layout,
    fingerprint,
    render_columns,
    score_sheet,
    self_check,
)
from app.services.budget.reserves import parse_reserve_schedule
from app.services.budget.workbook import _is_xls, load_workbook_any, normalize_to_xlsx

pytestmark = pytest.mark.unit


def _wb(sheets: dict[str, list[list]]):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, rows in sheets.items():
        ws = wb.create_sheet(title)
        for r in rows:
            ws.append(r)
    return wb


def _bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


BUDGET_ROWS = [
    ["", "2025 BUDGET", "2025 PROJECTED", "2026 BUDGET"],
    ["INCOME", None, None, None],
    ["   Assessments", 98000, 98000, 112677],
    ["   TOTAL INCOME", 98000, 98000, 112677],
    ["EXPENSES", None, None, None],
    ["ADMINISTRATION", None, None, None],
    ["   Management", 50000, 50000, 60000],
    ["   TOTAL ADMINISTRATION", 50000, 50000, 60000],
    ["UTILITIES", None, None, None],
    ["   Electricity", 48000, 48000, 52677],
    ["   TOTAL UTILITIES", 48000, 48000, 52677],
]


class TestSheetSelection:
    def test_scores_budget_sheet_above_summary_and_graph_tabs(self):
        wb = _wb(
            {
                "Totals": [["ADMIN", None, 60000], ["MAINT", None, 52677]],
                "2025": BUDGET_ROWS,
                "Graph #s": [["ADMIN", 0.5], ["UTIL", 0.5]],
            }
        )
        scores = {ws.title: score_sheet(ws) for ws in wb.worksheets}
        assert scores["2025"] > scores["Totals"]
        assert scores["2025"] > scores["Graph #s"]

    def test_build_layout_picks_the_budget_sheet(self):
        wb = _wb({"Totals": [["ADMIN", None, 1]] * 6, "2025": BUDGET_ROWS})
        assert build_layout(wb, 2027).sheet_title == "2025"


class TestColumnRoles:
    @pytest.mark.parametrize(
        "header,expected_role",
        [
            ("2026 BUDGET", "adopted"),
            ("2026 ADOPTED", "adopted"),
            ("2026 PROPOSED BUDGET", "adopted"),
            ("2025 PROJECTED", "projected"),
            ("Projected 2025", "projected"),
            ("2025 ACTUAL", "actual"),
        ],
    )
    def test_roles(self, header, expected_role):
        rows = [["", "2025 BUDGET", header], ["INCOME", None, None], ["   A", 1, 2]]
        wb = _wb({"Budget": rows})
        layout = build_layout(wb, 2027)
        role = next((v.role for v in layout.value_cols if v.header == header), None)
        assert role == expected_role

    def test_actual_column_is_never_the_prior_year_source(self):
        """Reading ACTUAL instead of ADOPTED was the silent MCP bug."""
        rows = [
            ["", "2025 ACTUAL", "2025 PROJECTED", "2026 ADOPTED"],
            ["INCOME", None, None, None],
            ["   Assessments", 1450, 922, 1250],
        ]
        layout = build_layout(_wb({"Budget": rows}), 2027)
        assert layout.prior_col == 4


class TestRenderColumns:
    def test_write_targets_are_distinct(self):
        """A single-adopted-column template must not write prior and proposed
        into the same column."""
        rows = [
            ["", "2025 ACTUAL", "2025 PROJECTED", "2026 ADOPTED"],
            ["INCOME", None, None, None],
            ["   Assessments", 1450, 922, 1250],
        ]
        layout = build_layout(_wb({"Budget": rows}), 2027)
        prior, projected, proposed, _notes = render_columns(layout, 2027)
        assert len({prior, projected, proposed}) == 3

    def test_shifts_forward_a_year_on_standard_template(self):
        layout = build_layout(_wb({"Budget": BUDGET_ROWS}), 2027)
        prior, _projected, proposed, _notes = render_columns(layout, 2027)
        # 2026 BUDGET (col 4) becomes the new proposed slot; 2025 (col 2) receives
        # last year's values.
        assert proposed == 4
        assert prior == 2


class TestSelfCheck:
    def test_balanced_budget_passes(self):
        wb = _wb({"Budget": BUDGET_ROWS})
        layout = build_layout(wb, 2027)
        from app.services.budget.stages.ingest import _parse_excel_budget

        lines, _ = _parse_excel_budget(_bytes(wb), 2027)
        assert self_check(wb[layout.sheet_title], layout, lines) == []

    def test_unbalanced_budget_is_flagged(self):
        rows = [
            ["", "2025 BUDGET", "2025 PROJECTED", "2026 BUDGET"],
            ["INCOME", None, None, None],
            ["   Assessments", 98000, 98000, 112677],
            ["EXPENSES", None, None, None],
            ["ADMINISTRATION", None, None, None],
            ["   Management", 1, 1, 5],  # nowhere near revenue
        ]
        wb = _wb({"Budget": rows})
        layout = build_layout(wb, 2027)
        from app.services.budget.stages.ingest import _parse_excel_budget

        lines, _ = _parse_excel_budget(_bytes(wb), 2027)
        problems = self_check(wb[layout.sheet_title], layout, lines)
        assert any("balance" in p.lower() for p in problems)

    def test_empty_lines_is_flagged(self):
        wb = _wb({"Budget": BUDGET_ROWS})
        layout = build_layout(wb, 2027)
        assert self_check(wb[layout.sheet_title], layout, []) != []


class TestFingerprint:
    def test_same_shape_different_amounts_matches(self):
        """Next year's file for the same template must reuse the confirmed layout."""

        def rows(scale: int):
            return [
                ["", "2025 BUDGET", "2025 PROJECTED", "2026 BUDGET"],
                ["INCOME", None, None, None],
                ["   Assessments", 100 * scale, 100 * scale, 120 * scale],
                ["EXPENSES", None, None, None],
                ["ADMINISTRATION", None, None, None],
                ["   Management", 100 * scale, 100 * scale, 120 * scale],
            ]

        a = _wb({"Budget": rows(1)})
        b = _wb({"Budget": rows(97)})
        la, lb = build_layout(a, 2027), build_layout(b, 2027)
        assert fingerprint(a["Budget"], la) == fingerprint(b["Budget"], lb)

    def test_line_count_does_not_change_the_signature(self):
        """Two associations on the same template carry different numbers of line
        items. If that changed the signature, one review per template would
        become one review per association — the whole point of the gate."""

        def template(extra: int):
            rows = [
                ["", "2025 BUDGET", "2025 PROJECTED", "2026 BUDGET"],
                ["INCOME", None, None, None],
                ["   Assessments", 100, 100, 120],
                ["   TOTAL INCOME", 100, 100, 120],
                ["EXPENSES", None, None, None],
                ["ADMINISTRATION", None, None, None],
            ]
            rows += [[f"   Item {i}", 50, 50, 60] for i in range(1 + extra)]
            rows.append(["   TOTAL ADMINISTRATION", 50, 50, 60])
            return rows

        a, b = _wb({"Budget": template(0)}), _wb({"Budget": template(11)})
        assert build_layout(a, 2027).signature == build_layout(b, 2027).signature

    def test_different_section_structure_differs(self):
        other = [
            ["", "2025 ACTUAL", "2025 PROJECTED", "2026 ADOPTED"],
            ["MEMBER ASSESSMENT", None, None, None],
            ["Member Assessments", 100, 100, 120],
            ["   TOTAL OPERATING", 100, 100, 120],
            ["BUILDING AND GROUNDS", None, None, None],
            ["Grounds", 100, 100, 120],
            ["   TOTAL BLDG & GROUNDS", 100, 100, 120],
        ]
        a, b = _wb({"Budget": BUDGET_ROWS}), _wb({"Budget": other})
        assert build_layout(a, 2027).signature != build_layout(b, 2027).signature


class TestSubtotalRows:
    def test_duplicate_total_labels_map_to_separate_rows(self):
        """RVL labels BOTH its revenue and its expense subtotal "TOTAL OPERATING".
        Matching by text cannot tell them apart, so subtotal rows are recorded
        per group and written by row index."""
        rows = [
            ["", "2025 BUDGET", "2025 PROJECTED", "2026 BUDGET"],
            ["REVENUES:", None, None, None],
            ["OPERATING", None, None, None],
            ["   Assessments", 98000, 98000, 112677],
            ["   TOTAL OPERATING", 98000, 98000, 112677],
            ["EXPENSES FOR THE HOA", None, None, None],
            ["Operating", None, None, None],
            ["   Management", 98000, 98000, 112677],
            ["   TOTAL OPERATING", 98000, 98000, 112677],
        ]
        layout = build_layout(_wb({"Budget": rows}), 2027)

        assert layout.subtotal_rows["REVENUE_OPERATING::OPERATING"] == 5
        assert layout.subtotal_rows["OTHER::Operating"] == 9

    def test_split_section_keeps_a_row_per_heading(self):
        """MCP splits BUILDING AND GROUNDS from MAINTENANCE; both normalize to
        MAINTENANCE but each keeps its own printed subtotal row."""
        rows = [
            ["", "2025 ACTUAL", "2025 PROJECTED", "2026 ADOPTED"],
            ["MEMBER ASSESSMENT", None, None, None],
            ["Member Assessments", 100, 100, 100],
            ["   TOTAL OPERATING", 100, 100, 100],
            ["BUILDING AND GROUNDS", None, None, None],
            ["Grounds Maintenance", 60, 60, 60],
            ["   TOTAL BLDG & GROUNDS", 60, 60, 60],
            ["MAINTENANCE", None, None, None],
            ["Misc Repair", 40, 40, 40],
            ["   TOTAL MAINTENANCE", 40, 40, 40],
        ]
        layout = build_layout(_wb({"Budget": rows}), 2027)

        assert layout.subtotal_rows["MAINTENANCE::BUILDING AND GROUNDS"] == 7
        assert layout.subtotal_rows["MAINTENANCE::MAINTENANCE"] == 10


class TestTitleBlock:
    """Every rendered budget carries the same three-line heading:

    SARABAY WOODS HOA
    2026 APPROVED OPERATING BUDGET
    JANUARY 1, 2026 - DECEMBER 31, 2026
    """

    def _render(self, rows, name="Sarabay Woods HOA", year=2027):
        from app.services.budget.schema import IngestedLine, IngestResult
        from app.services.budget.stages import assemble, project, render, validate
        from app.services.budget.stages.ingest import _parse_excel_budget

        raw = _bytes(_wb({"Budget": rows}))
        lines, layout = _parse_excel_budget(raw, year)
        ing = IngestResult(
            months_elapsed=6,
            lines=[
                IngestedLine(
                    label=x["label"],
                    section=x["section"],
                    prior_year=x["prior_year"],
                    ytd_actual=(x["prior_year"] or 0) / 2,
                    source_section=x.get("source_section"),
                )
                for x in lines
            ],
            subtotal_rows=layout.subtotal_rows,
        )
        projected, review_labels = project.run(ing)
        budget = assemble.run(
            ingest=ing,
            projected=projected,
            review_labels=review_labels,
            association_name=name,
            budget_year=year,
            prior_reserve_schedule=None,
        )
        out = render.run(budget, validate.run(budget), raw)
        return openpyxl.load_workbook(io.BytesIO(out.xlsx_bytes))["Budget"]

    def _title_at(self, ws, col=1):
        for r in range(1, 12):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, str) and v.strip().endswith("HOA"):
                return [str(ws.cell(row=r + i, column=col).value or "").strip() for i in range(3)]
        return []

    def test_added_when_the_sheet_has_no_title(self):
        """Headers on row 1 leave no room — rows must be inserted."""
        ws = self._render(BUDGET_ROWS)
        assert self._title_at(ws) == [
            "SARABAY WOODS HOA",
            "2027 APPROVED OPERATING BUDGET",
            "JANUARY 1, 2027 - DECEMBER 31, 2027",
        ]

    def test_existing_title_is_rewritten_in_place(self):
        """SBW already carries a title; it is updated, not duplicated."""
        rows = [
            ["OLD NAME HOA", None, None, None],
            ["2026 APPROVED OPERATING BUDGET", None, None, None],
            ["JANUARY 1, 2026 - DECEMBER 31, 2026", None, None, None],
            [None, None, None, None],
            ["", "2025 BUDGET", "2025 PROJECTED", "2026 BUDGET"],
            ["INCOME", None, None, None],
            ["   Assessments", 100, 100, 120],
            ["   TOTAL INCOME", 100, 100, 120],
            ["EXPENSES", None, None, None],
            ["ADMINISTRATION", None, None, None],
            ["   Management", 100, 100, 120],
            ["   TOTAL ADMINISTRATION", 100, 100, 120],
        ]
        ws = self._render(rows)
        assert self._title_at(ws) == [
            "SARABAY WOODS HOA",
            "2027 APPROVED OPERATING BUDGET",
            "JANUARY 1, 2027 - DECEMBER 31, 2027",
        ]
        # Rewritten, not appended: the superseded year appears nowhere.
        text = [
            str(ws.cell(row=r, column=1).value or "") for r in range(1, min(ws.max_row, 12) + 1)
        ]
        assert not any("2026 APPROVED" in t for t in text)

    def test_subtotal_formulas_survive_the_row_shift(self):
        """Inserting title rows must not leave =SUM formulas pointing at the
        wrong rows — openpyxl does not rewrite references when rows move."""
        ws = self._render(BUDGET_ROWS)
        totals = [
            ws.cell(row=r, column=2).value
            for r in range(1, ws.max_row + 1)
            if str(ws.cell(row=r, column=1).value or "").strip().upper().startswith("TOTAL")
        ]
        assert totals, "no TOTAL rows found"
        assert all(isinstance(v, str) and v.startswith("=") for v in totals), totals


class TestColumnHeaders:
    def _render(self, rows, year=2027):
        from app.services.budget.schema import IngestedLine, IngestResult
        from app.services.budget.stages import assemble, project, render, validate
        from app.services.budget.stages.ingest import _parse_excel_budget

        raw = _bytes(_wb({"Budget": rows}))
        lines, layout = _parse_excel_budget(raw, year)
        ing = IngestResult(
            months_elapsed=6,
            lines=[
                IngestedLine(
                    label=x["label"],
                    section=x["section"],
                    prior_year=x["prior_year"],
                    ytd_actual=(x["prior_year"] or 0) / 2,
                    source_section=x.get("source_section"),
                )
                for x in lines
            ],
            subtotal_rows=layout.subtotal_rows,
        )
        projected, review_labels = project.run(ing)
        budget = assemble.run(
            ingest=ing,
            projected=projected,
            review_labels=review_labels,
            association_name="Test HOA",
            budget_year=year,
            prior_reserve_schedule=None,
        )
        out = render.run(budget, validate.run(budget), raw)
        wb2 = openpyxl.load_workbook(io.BytesIO(out.xlsx_bytes))["Budget"]
        return wb2, build_layout(openpyxl.load_workbook(io.BytesIO(out.xlsx_bytes)), year)

    def test_all_three_headers_roll_forward_keeping_their_wording(self):
        """Associations word these headers their own way. Only the year changes."""
        rows = [
            ["", "2025 ACTUAL", "2025 PROJECTED", "2026 ADOPTED"],
            ["INCOME", None, None, None],
            ["   Assessments", 100, 100, 120],
            ["   TOTAL INCOME", 100, 100, 120],
            ["EXPENSES", None, None, None],
            ["ADMINISTRATION", None, None, None],
            ["   Management", 100, 100, 120],
            ["   TOTAL ADMINISTRATION", 100, 100, 120],
        ]
        ws, layout = self._render(rows)
        headers = [ws.cell(row=layout.header_row, column=c).value for c in (2, 3, 4)]
        assert headers == ["2026 ACTUAL", "2026 PROJECTED", "2027 ADOPTED"]

    def test_proposed_header_names_the_new_year_but_values_are_untouched(self):
        """The manager fills this column in; the heading must still say which
        year they are filling in for."""
        rows = [
            ["", "2025 BUDGET", "2025 PROJECTED", "2026 BUDGET"],
            ["INCOME", None, None, None],
            ["   Assessments", 100, 100, 555],
            ["   TOTAL INCOME", 100, 100, 555],
            ["EXPENSES", None, None, None],
            ["ADMINISTRATION", None, None, None],
            ["   Management", 100, 100, 555],
            ["   TOTAL ADMINISTRATION", 100, 100, 555],
        ]
        ws, layout = self._render(rows)
        assert ws.cell(row=layout.header_row, column=4).value == "2027 BUDGET"

        # The manager's figures survive.
        values = [
            ws.cell(row=r, column=4).value
            for r in range(layout.header_row + 1, ws.max_row + 1)
            if str(ws.cell(row=r, column=1).value or "").strip() == "Assessments"
        ]
        assert values == [555]

    def test_header_without_a_year_falls_back_to_a_default(self):
        rows = [
            ["", "BUDGET", "PROJECTED", "PROPOSED"],
            ["INCOME", None, None, None],
            ["   Assessments", 100, 100, 120],
            ["   TOTAL INCOME", 100, 100, 120],
            ["EXPENSES", None, None, None],
            ["ADMINISTRATION", None, None, None],
            ["   Management", 100, 100, 120],
            ["   TOTAL ADMINISTRATION", 100, 100, 120],
        ]
        ws, layout = self._render(rows)
        assert "2026" in str(ws.cell(row=layout.header_row, column=2).value)


class TestGlAccounts:
    def test_gl_codes_are_read_and_sent_to_the_extractor(self):
        """The workbook and the report word lines differently ("Assessments" vs
        "Member Assessments") but agree on the GL account, so the account number
        has to reach the model."""
        from app.services.budget.stages.ingest import _format_line_list, _parse_excel_budget

        rows = [
            ["", "", "2025 BUDGET", "2025 PROJECTED", "2026 BUDGET"],
            ["", "INCOME", None, None, None],
            [5000, "   Assessments", 98000, 98000, 112677],
            [1030, "   Oper Interest Income", 900, 900, 2000],
            ["", "   TOTAL INCOME", 98900, 98900, 114677],
            ["", "EXPENSES", None, None, None],
            ["", "ADMINISTRATION", None, None, None],
            [6050, "   Management Fees", 98900, 98900, 114677],
            ["", "   TOTAL ADMINISTRATION", 98900, 98900, 114677],
        ]
        lines, layout = _parse_excel_budget(_bytes(_wb({"Budget": rows})), budget_year=2027)

        assert layout.gl_col == 1
        by_label = {ln["label"]: ln for ln in lines}
        assert by_label["Assessments"]["gl_account"] == "5000"
        assert by_label["Management Fees"]["gl_account"] == "6050"

        prompt = _format_line_list(lines)
        assert "[5000] Assessments" in prompt
        assert "GL ACCOUNT" in prompt.upper()

    def test_workbook_without_gl_column_still_parses(self):
        from app.services.budget.stages.ingest import _parse_excel_budget

        lines, layout = _parse_excel_budget(_bytes(_wb({"Budget": BUDGET_ROWS})), budget_year=2027)
        assert layout.gl_col is None
        assert lines
        assert all(ln["gl_account"] is None for ln in lines)


class TestUnmatchedLines:
    def test_line_with_no_ytd_is_flagged_and_highlighted(self):
        """A line the financial report has no match for renders a blank projected
        cell. Silently blank is how it used to look; it must be visible."""
        from app.services.budget.schema import IngestedLine, IngestResult
        from app.services.budget.stages import assemble, project, render, validate
        from app.services.budget.stages.ingest import _parse_excel_budget

        wb = _wb({"Budget": BUDGET_ROWS})
        raw = _bytes(wb)
        lines, layout = _parse_excel_budget(raw, 2027)

        ing = IngestResult(
            months_elapsed=6,
            lines=[
                IngestedLine(
                    label=x["label"],
                    section=x["section"],
                    prior_year=x["prior_year"],
                    # "Management" has no match in the report.
                    ytd_actual=None if x["label"] == "Management" else 100.0,
                    source_section=x.get("source_section"),
                )
                for x in lines
            ],
            subtotal_rows=layout.subtotal_rows,
        )
        projected, review_labels = project.run(ing)
        budget = assemble.run(
            ingest=ing,
            projected=projected,
            review_labels=review_labels,
            association_name="T",
            budget_year=2027,
            prior_reserve_schedule=None,
        )
        out = render.run(budget, validate.run(budget), raw)

        flags = [f for f in out.review_flags if "no matching line" in f]
        assert any("Management" in f for f in flags), out.review_flags

        ws = openpyxl.load_workbook(io.BytesIO(out.xlsx_bytes))[layout.sheet_title]
        for r in range(layout.data_start_row, ws.max_row + 1):
            if str(ws.cell(row=r, column=layout.label_col).value or "").strip() == "Management":
                cell = ws.cell(row=r, column=layout.label_col)
                assert cell.fill.fill_type == "solid"
                assert cell.fill.start_color.rgb.endswith("FFF0CC")
                break
        else:
            raise AssertionError("Management row not found in rendered sheet")


class TestReserveSchedule:
    def test_parses_stacked_header_and_skips_shortfall_column(self):
        """ "Balance Needed" is a shortfall, not the current balance."""
        wb = _wb(
            {
                "Budget": BUDGET_ROWS,
                "Reserve": [
                    ["", "Total", "Remain", "Estimated", "Balance", "Current", "", "Annual"],
                    ["", "Life", "Life", "Total Cost", "Needed", "Balance", "", "Funding"],
                    ["Roof", 15, 9, 875000, 380195.75, 494804.25, None, 42243.97],
                    ["Clubhouse", 15, 8, 65000, 43140.23, 21859.77, None, 5392.53],
                ],
            }
        )
        layout = build_layout(wb, 2027)
        items = parse_reserve_schedule(wb, layout)

        assert len(items) == 2
        roof = items[0]
        assert roof["label"] == "Roof"
        assert roof["total_life_years"] == 15
        assert roof["remaining_life_years"] == 9
        assert roof["replacement_cost"] == 875000
        # Current Balance, not Balance Needed.
        assert roof["current_balance"] == 494804.25
        assert roof["required_deposit"] == pytest.approx(42243.97)

    def test_grand_total_row_is_excluded(self):
        wb = _wb(
            {
                "Budget": BUDGET_ROWS,
                "Reserves": [
                    ["", "Total Life", "Remaining", "Estimated Current", "Balance at 12/2025"],
                    ["", "In Years", "Life In Years", "Replacement Cost", ""],
                    ["Total Paving", 18, 14, 145000, 75000],
                    ["Total Reserves", None, None, 1129567, None],
                ],
            }
        )
        labels = [i["label"] for i in parse_reserve_schedule(wb, build_layout(wb, 2027))]
        # "Total Paving" is a real component; "Total Reserves" is the grand total.
        assert labels == ["Total Paving"]

    def test_no_reserve_sheet_returns_empty(self):
        wb = _wb({"Budget": BUDGET_ROWS})
        assert parse_reserve_schedule(wb, build_layout(wb, 2027)) == []


class TestWorkbookCompat:
    def test_xlsx_passes_through_untouched(self):
        raw = _bytes(_wb({"Budget": BUDGET_ROWS}))
        assert normalize_to_xlsx(raw, "x.xlsx") is raw

    def test_detects_xls_by_magic_number(self):
        assert _is_xls(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1rest", "mislabelled.xlsx")
        assert not _is_xls(b"PK\x03\x04rest", "real.xlsx")

    def test_load_workbook_any_reads_xlsx(self):
        wb = load_workbook_any(_bytes(_wb({"Budget": BUDGET_ROWS})), "b.xlsx")
        assert "Budget" in wb.sheetnames
