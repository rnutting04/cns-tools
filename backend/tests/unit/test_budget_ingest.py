"""Unit tests for the budget Excel ingest parser (in-memory xlsx; no DB, no AI)."""

import io

import openpyxl
import pytest

from app.services.budget.schema import BudgetSection
from app.services.budget.stages.ingest import _parse_excel_budget, _to_number

pytestmark = pytest.mark.unit


def _parse(xlsx_bytes: bytes, budget_year: int) -> list[dict]:
    """_parse_excel_budget returns (lines, layout); most tests only want lines."""
    lines, _layout = _parse_excel_budget(xlsx_bytes, budget_year)
    return lines


def _xlsx(rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget"
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# A standard-format sheet. `amt` lets us render the value columns either as real
# numbers or as text ("998,385"), which is how some association templates store
# them — the case that previously produced "No line items found".
def _standard_rows(amt) -> list[list]:
    return [
        [None, "2025 BUDGET", "2025 Projected", "2026 BUDGET"],
        ["INCOME", None, None, None],
        ["OPERATING", None, None, None],
        ["Assessments", amt(998385), amt(998385), amt(1345442)],
        ["Other Income", amt(500), amt(1500), amt(0)],
        ["   TOTAL OPERATING", amt(998885), amt(999885), amt(1345442)],
        ["RESERVES", None, None, None],
        ["Reserve Assessments", amt(137665), amt(137665), amt(256480)],
        ["   TOTAL RESERVES", amt(137665), amt(137665), amt(256480)],
        ["TOTAL INCOME", amt(1136550), amt(1137550), amt(1601922)],
        ["EXPENSES", None, None, None],
        ["ADMINISTRATION", None, None, None],
        ["Management Fees", amt(33930), amt(33930), amt(35287)],
        ["CPA", amt(5325), amt(5325), amt(5500)],
        ["   TOTAL ADMINISTRATION", amt(39255), amt(39255), amt(40787)],
        ["UTILITIES", None, None, None],
        ["Water/Sewer", amt(176000), amt(214013), amt(221503)],
        ["   TOTAL UTILITIES", amt(176000), amt(214013), amt(221503)],
        ["RESERVES", None, None, None],
        ["Roof", amt(56097), amt(56097), amt(42244)],
        ["TOTAL RESERVES", amt(56097), amt(56097), amt(42244)],
    ]


class TestToNumber:
    @pytest.mark.parametrize(
        "value,expected",
        [
            (1000, 1000.0),
            (1234.5, 1234.5),
            (" 998,385 ", 998385.0),
            ("$1,000.00", 1000.0),
            ("(1,200)", -1200.0),  # accounting-style negative
            ("0", 0.0),
        ],
    )
    def test_parses(self, value, expected):
        assert _to_number(value) == expected

    @pytest.mark.parametrize("value", [None, "", "  ", "-", "6810a", "n/a", True, False])
    def test_non_numeric(self, value):
        assert _to_number(value) is None


class TestParseExcelBudget:
    def test_text_formatted_numbers_are_parsed(self):
        """Regression: text-stored amounts must not produce an empty parse."""
        xb = _xlsx(_standard_rows(lambda x: f" {x:,} "))
        lines = _parse(xb, budget_year=2027)
        assert lines, "text-formatted amounts should still yield line items"

    def test_text_and_numeric_agree(self):
        text = _parse(_xlsx(_standard_rows(lambda x: f"{x:,}")), budget_year=2027)
        numeric = _parse(_xlsx(_standard_rows(float)), budget_year=2027)
        assert [(ln["label"], ln["section"], ln["prior_year"]) for ln in text] == [
            (ln["label"], ln["section"], ln["prior_year"]) for ln in numeric
        ]

    def test_sections_and_prior_year_column(self):
        # budget_year 2027 -> prior year is 2026 -> read the "2026 BUDGET" column.
        lines = _parse(_xlsx(_standard_rows(lambda x: f" {x:,} ")), budget_year=2027)
        by_label = {ln["label"]: ln for ln in lines}

        assert by_label["Assessments"]["section"] == BudgetSection.REVENUE_OPERATING
        assert by_label["Reserve Assessments"]["section"] == BudgetSection.REVENUE_RESERVES
        assert by_label["Management Fees"]["section"] == BudgetSection.ADMINISTRATION
        assert by_label["Water/Sewer"]["section"] == BudgetSection.UTILITIES
        assert by_label["Roof"]["section"] == BudgetSection.RESERVES

        # prior_year must come from 2026 BUDGET, not 2025.
        assert by_label["Assessments"]["prior_year"] == 1345442.0
        assert by_label["Management Fees"]["prior_year"] == 35287.0

        # TOTAL / subtotal rows are excluded.
        assert not [ln for ln in lines if ln["label"].upper().startswith("TOTAL")]


# ── Regression cases from real association workbooks ─────────────────────────
# Each of these reproduces a layout that silently mis-parsed or hard-failed
# before layout.py centralized structure detection. They use synthetic sheets
# rather than the files in test_data/ so the suite stays hermetic.


class TestRealWorldLayouts:
    def test_budget_sheet_found_when_not_named_budget(self):
        """GOR: budget lives on a sheet named "2025"; "Totals" sorts ahead of it.

        The old name-match picked the 6-row Totals tab and produced zero lines.
        """
        wb = openpyxl.Workbook()
        totals = wb.active
        totals.title = "Totals"
        for r in [
            ["ADMIN", None, 155587],
            ["MAINT", None, 232645],
            ["OTHER", None, 561810],
            ["UTILITIES", None, 423800],
            ["RESERVE", None, 256479],
            ["GROUNDS", None, 99600],
        ]:
            totals.append(r)

        real = wb.create_sheet("2025")
        for r in _standard_rows(float):
            real.append(r)

        buf = io.BytesIO()
        wb.save(buf)
        lines, layout = _parse_excel_budget(buf.getvalue(), budget_year=2027)

        assert layout.sheet_title == "2025"
        assert lines
        assert {ln["label"] for ln in lines} >= {"Assessments", "Management Fees", "Water/Sewer"}

    def test_adopted_column_without_the_word_budget(self):
        """MCP: columns are ACTUAL / PROJECTED / ADOPTED — none says "BUDGET".

        The old detector required the literal word "budget", found nothing, and
        fell back to label_col+1 — the ACTUAL column — silently.
        """
        rows = [
            ["", "2025 ACTUAL", "2025 PROJECTED", "2026 ADOPTED"],
            ["MEMBER ASSESSMENT", None, None, None],
            ["Member Assessments", 161477.44, 161477.44, 161479.08],
            ["   TOTAL OPERATING", 161477.44, 161477.44, 161479.08],
            ["ADMINISTRATION", None, None, None],
            ["Postage & Printing", 1450, 922.53, 1250],
            ["Management Fees", 12940, 12940, 13324.08],
            ["   TOTAL ADMINISTRATION", 14390, 13862.53, 14574.08],
            ["UTILITIES", None, None, None],
            ["Electricity", 15000, 14005, 15000],
            ["   TOTAL UTILITIES", 15000, 14005, 15000],
        ]
        lines, layout = _parse_excel_budget(_xlsx(rows), budget_year=2027)
        by_label = {ln["label"]: ln for ln in lines}

        # prior_year must come from "2026 ADOPTED" (1250), never "2025 ACTUAL" (1450).
        assert by_label["Postage & Printing"]["prior_year"] == 1250
        assert layout.prior_col == 4

    def test_member_assessment_header_is_operating_revenue(self):
        """MCP: "MEMBER ASSESSMENT" matched no income keyword, so the main
        revenue line was dropped and every expense fell into REVENUE_RESERVES."""
        rows = [
            ["", "2025 ACTUAL", "2025 PROJECTED", "2026 ADOPTED"],
            ["MEMBER ASSESSMENT", None, None, None],
            ["Member Assessments", 161477.44, 161477.44, 161479.08],
            ["   TOTAL OPERATING", 161477.44, 161477.44, 161479.08],
            ["ADMINISTRATION", None, None, None],
            ["Management Fees", 12940, 12940, 13324.08],
            ["   TOTAL ADMINISTRATION", 12940, 12940, 13324.08],
            ["BUILDING AND GROUNDS", None, None, None],
            ["Grounds Maintenance Contract", 51500, 57144, 58858],
            ["   TOTAL BLDG & GROUNDS", 51500, 57144, 58858],
        ]
        lines, _ = _parse_excel_budget(_xlsx(rows), budget_year=2027)
        by_label = {ln["label"]: ln for ln in lines}

        assert by_label["Member Assessments"]["section"] == BudgetSection.REVENUE_OPERATING
        assert by_label["Management Fees"]["section"] == BudgetSection.ADMINISTRATION
        assert by_label["Grounds Maintenance Contract"]["section"] == BudgetSection.MAINTENANCE
        # Not everything in one bucket.
        assert len({ln["section"] for ln in lines}) >= 3

    def test_flat_expense_block_is_not_revenue(self):
        """RVL: expenses sit under a bare "Operating" header below "EXPENSES FOR
        THE HOA". The old parser left every one of them tagged as revenue."""
        rows = [
            ["", "2025 BUDGET", "2025 PROJECTED", "2026 BUDGET"],
            ["REVENUES:", None, None, None],
            ["OPERATING", None, None, None],
            ["   Assessments", 98000, 98000, 112677],
            ["   TOTAL OPERATING", 98000, 98000, 112677],
            ["EXPENSES FOR THE HOA", None, None, None],
            ["Operating", None, None, None],
            ["   Management", 9600, 9600, 10080],
            ["   Lawn Care Contract", 35400, 33840, 35400],
            ["   TOTAL OPERATING", 45000, 43440, 45480],
        ]
        lines, _ = _parse_excel_budget(_xlsx(rows), budget_year=2027)
        by_label = {ln["label"]: ln for ln in lines}

        assert by_label["Assessments"]["section"] == BudgetSection.REVENUE_OPERATING
        for expense in ("Management", "Lawn Care Contract"):
            assert not by_label[expense]["section"].value.startswith(
                "REVENUE"
            ), f"{expense} was booked as revenue"

    def test_assessment_rate_preamble_is_not_ingested(self):
        """PRD: 25 rows of per-unit rates by building type sit above "INCOME".
        They carry numbers but are not budget line items."""
        rows = [
            ["", "2025 BUDGET", "2025 PROJECTED", "2026 BUDGET"],
            ["Quarterly Assessment", None, None, None],
            ["Sago", 2390.57, 2390.57, 2841.91],
            ["Norfork", 3091.81, 3091.81, 3675.53],
            ["Savannah", 5498.32, 5498.32, 6536.39],
            ["INCOME", None, None, None],
            ["   OPERATING", None, None, None],
            ["   Member Assessments", 572048.32, 583172.00, 583343.14],
            ["   TOTAL OPERATING", 572048.32, 583172.00, 583343.14],
            ["EXPENSES", None, None, None],
            ["   ADMINISTRATION", None, None, None],
            ["   Management Fees", 10036.32, 10036.00, 10538.14],
            ["   TOTAL ADMINISTRATION", 10036.32, 10036.00, 10538.14],
        ]
        lines, _ = _parse_excel_budget(_xlsx(rows), budget_year=2027)
        labels = {ln["label"] for ln in lines}

        assert "Member Assessments" in labels
        for rate_row in ("Sago", "Norfork", "Savannah"):
            assert rate_row not in labels, f"{rate_row} rate row leaked in as a line item"

    def test_self_check_flags_everything_in_one_section(self):
        """The confidence gate must catch the exact shape of the old silent bug."""
        from app.services.budget.layout import build_layout, self_check
        from app.services.budget.workbook import load_workbook_any

        rows = [
            ["", "2025 BUDGET", "2025 PROJECTED", "2026 BUDGET"],
            ["REVENUES:", None, None, None],
            ["   Assessments", 98000, 98000, 112677],
            ["   TOTAL OPERATING", 98000, 98000, 112677],
        ]
        xb = _xlsx(rows)
        wb = load_workbook_any(xb)
        layout = build_layout(wb, 2027)
        lines, _ = _parse_excel_budget(xb, budget_year=2027)

        problems = self_check(wb[layout.sheet_title], layout, lines)
        assert any("expense" in p.lower() for p in problems)
